import openpyxl

wbObj = openpyxl.load_workbook("C:\\Users\\dines\\OneDrive\\Documents\\MyPythonProject\\DataSetUp.xlsx")
wsObj = wbObj.active
# cellObj = wsObj.cell(row=5, column=5)
# cellObj.value = "hello"
# wsObj['F5'] = "janani"
# print(wsObj['F5'].value)
# cellObj = wsObj.cell(row=5, column=5, value="hello")
# print(cellObj.value)
wbObj.save("C:\\Users\\dines\\OneDrive\\Documents\\MyPythonProject\\DataSetUp.xlsx")
print(wsObj.max_row, wsObj.max_column)
for i in range(1,wsObj.max_row+1):
    if wsObj.cell(row=i, column=1).value == 4:
        for j in range(1,wsObj.max_column+1):
            # print("j=",j)
            print(wsObj.cell(row = i, column = j).value)
